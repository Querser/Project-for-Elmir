from __future__ import annotations

from datetime import datetime, timezone
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import and_, func, or_
from sqlalchemy.orm import Session, selectinload

from app.models.ban import Ban
from app.models.debt import Debt, DebtStatus
from app.models.enrollment import Enrollment, EnrollmentStatus
from app.models.payment import Payment, PaymentStatus
from app.models.training import Training
from app.models.user import User
from app.services.payment_retention_service import purge_payments_older_than_quarter, quarter_cutoff


def _value_to_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        if value.tzinfo is None:
            return value.replace(tzinfo=timezone.utc).isoformat()
        return value.astimezone(timezone.utc).isoformat()
    if isinstance(value, Decimal):
        return f"{value:.2f}"
    enum_value = getattr(value, "value", None)
    if enum_value is not None:
        return str(enum_value)
    return str(value)


def _auto_fit_columns(ws) -> None:
    for column_cells in ws.columns:
        length = 0
        col_letter = column_cells[0].column_letter
        for cell in column_cells:
            cell_value = "" if cell.value is None else str(cell.value)
            length = max(length, len(cell_value))
        ws.column_dimensions[col_letter].width = min(max(12, length + 2), 60)


def _workbook_to_bytes(workbook: Workbook) -> bytes:
    bio = BytesIO()
    workbook.save(bio)
    return bio.getvalue()


def build_users_export_xlsx(db: Session) -> bytes:
    now = datetime.now(timezone.utc)
    users = (
        db.query(User)
        .options(selectinload(User.level))
        .order_by(User.id.asc())
        .all()
    )
    user_ids = [u.id for u in users]

    debts_map: dict[int, int] = {}
    if user_ids:
        debt_rows = (
            db.query(Debt.user_id, func.count(Debt.id))
            .filter(Debt.user_id.in_(user_ids), Debt.status == DebtStatus.OPEN)
            .group_by(Debt.user_id)
            .all()
        )
        debts_map = {int(uid): int(cnt or 0) for uid, cnt in debt_rows}

    active_ban_map: dict[int, Ban] = {}
    if user_ids:
        active_bans = (
            db.query(Ban)
            .filter(Ban.user_id.in_(user_ids))
            .filter(
                and_(
                    Ban.active.is_(True),
                    or_(Ban.until.is_(None), Ban.until >= now),
                )
            )
            .order_by(Ban.user_id.asc(), Ban.created_at.desc())
            .all()
        )
        for ban in active_bans:
            uid = int(ban.user_id)
            if uid not in active_ban_map:
                active_ban_map[uid] = ban

    wb = Workbook()
    ws = wb.active
    ws.title = "Users"

    headers = [
        "id",
        "telegram_id",
        "username",
        "first_name",
        "last_name",
        "phone",
        "gender",
        "birth_date",
        "level_id",
        "level_name",
        "rating",
        "cups",
        "is_telegram_public",
        "payer_id",
        "card_last4",
        "avatar_url",
        "is_active",
        "is_admin",
        "created_at",
        "updated_at",
        "has_active_ban",
        "active_ban_type",
        "active_ban_reason",
        "active_ban_until",
        "open_debts_count",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for user in users:
        active_ban = active_ban_map.get(int(user.id))
        ws.append(
            [
                int(user.id),
                int(user.telegram_id) if user.telegram_id is not None else "",
                user.username or "",
                user.first_name or "",
                user.last_name or "",
                user.phone or "",
                user.gender or "",
                _value_to_text(user.birth_date),
                user.level_id if user.level_id is not None else "",
                getattr(user.level, "name", "") if getattr(user, "level", None) else "",
                int(user.rating or 0),
                int(user.cups or 0),
                bool(getattr(user, "is_telegram_public", True)),
                user.payer_id or "",
                user.card_last4 or "",
                user.avatar_url or "",
                bool(user.is_active),
                bool(user.is_admin),
                _value_to_text(user.created_at),
                _value_to_text(user.updated_at),
                bool(active_ban is not None),
                _value_to_text(getattr(active_ban, "type", "")),
                _value_to_text(getattr(active_ban, "reason", "")),
                _value_to_text(getattr(active_ban, "until", None)),
                int(debts_map.get(int(user.id), 0)),
            ]
        )

    _auto_fit_columns(ws)
    return _workbook_to_bytes(wb)


def build_payments_export_last_quarter_xlsx(db: Session) -> tuple[bytes, int]:
    deleted = purge_payments_older_than_quarter(db)
    cutoff = quarter_cutoff()

    payments = (
        db.query(Payment, User, Training)
        .outerjoin(User, User.id == Payment.user_id)
        .outerjoin(Training, Training.id == Payment.training_id)
        .filter(Payment.created_at >= cutoff)
        .order_by(Payment.created_at.desc(), Payment.id.desc())
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Payments"

    headers = [
        "payment_id",
        "provider_payment_id",
        "status",
        "is_success",
        "amount",
        "currency",
        "created_at",
        "paid_at",
        "user_id",
        "telegram_id",
        "username",
        "first_name",
        "last_name",
        "training_id",
        "training_title",
        "training_start_at",
        "period_start_utc",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for payment, user, training in payments:
        status_text = _value_to_text(payment.status).lower()
        is_success = status_text in {
            PaymentStatus.PAID.value,
            PaymentStatus.REFUNDED.value,
        }
        ws.append(
            [
                int(payment.id),
                payment.provider_payment_id or "",
                status_text,
                bool(is_success),
                _value_to_text(payment.amount),
                payment.currency or "RUB",
                _value_to_text(payment.created_at),
                _value_to_text(payment.paid_at),
                int(payment.user_id) if payment.user_id is not None else "",
                int(user.telegram_id) if user and user.telegram_id is not None else "",
                user.username if user else "",
                user.first_name if user else "",
                user.last_name if user else "",
                int(payment.training_id) if payment.training_id is not None else "",
                getattr(training, "title", "") if training else "",
                _value_to_text(getattr(training, "start_at", None) if training else None),
                _value_to_text(cutoff),
            ]
        )

    _auto_fit_columns(ws)
    return _workbook_to_bytes(wb), deleted


def build_training_participants_export_xlsx(db: Session, *, training_id: int) -> tuple[bytes, int]:
    training = db.query(Training).filter(Training.id == training_id).first()
    if training is None:
        raise ValueError("training_not_found")

    rows = (
        db.query(User, Enrollment)
        .join(Enrollment, Enrollment.user_id == User.id)
        .filter(Enrollment.training_id == training_id)
        .filter(Enrollment.status.in_([EnrollmentStatus.ACTIVE, EnrollmentStatus.RESERVE]))
        .order_by(Enrollment.is_reserve.asc(), User.last_name.asc(), User.first_name.asc(), User.id.asc())
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Participants"

    headers = [
        "phone",
        "first_name",
        "last_name",
        "telegram_username",
        "id",
        "telegram_id",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for user, _enrollment in rows:
        ws.append(
            [
                user.phone or "",
                user.first_name or "",
                user.last_name or "",
                user.username or "",
                int(user.id),
                int(user.telegram_id) if user.telegram_id is not None else "",
            ]
        )

    _auto_fit_columns(ws)
    return _workbook_to_bytes(wb), len(rows)
